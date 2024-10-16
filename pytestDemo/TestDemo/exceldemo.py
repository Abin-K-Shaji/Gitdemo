import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Abin K Shaji\\Downloads\\Defect Report Generation.xlsx")
sheet=book.active
Dict={}
cell=sheet.cell(row=3,column=2)
print(cell.value)
for i in range(3,sheet.max_row+1):
    if  sheet.cell(row=i,column=1).value=="DEF_02":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=3,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)
"//div[text()='Apple']/parent::div/parent::div/div[4]"